import os
from asyncio.log import logger
from openpyxl import Workbook
from loguru import logger
def save_weather_in_exel(weather: dict[dict], city: str):
    """write data weather in excel file"""
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Weather"
    row = 2
    sheet.append(["Дата прогноза", "Ср. температура", "Мин. температура утром", "Макс. температура утром", 
                "Минимальная температура днём", "Макс. температура днём", "Мин. температура вечером", 
                "Макс. темпрература вечером", "Мин. температура ночью", "Макс. температура ночью", 
                "Атмосферное давление", "Влажность утром", "Влажность днём", "Влажность вечером", 
                "Влажность ночью", "Погода утром", "Погода днём", "Погода вечером", "Погода ночью",
                ])

    for day in weather:
        sheet.cell(row=row, column=1).value = day.get('forecast_date')
        sheet.cell(row=row, column=2).value = day.get('average_temperature')
        sheet.cell(row=row, column=3).value = day.get('max_min_temperature').get('min_temperature_morning')
        sheet.cell(row=row, column=4).value = day.get('max_min_temperature').get('max_temperature_morning')
        sheet.cell(row=row, column=5).value = day.get('max_min_temperature').get('min_temperature_day')
        sheet.cell(row=row, column=6).value = day.get('max_min_temperature').get('max_temperature_day')
        sheet.cell(row=row, column=7).value = day.get('max_min_temperature').get('min_temperature_evening')
        sheet.cell(row=row, column=8).value = day.get('max_min_temperature').get('max_temperature_evening')
        sheet.cell(row=row, column=9).value = day.get('max_min_temperature').get('min_temperature_night')
        sheet.cell(row=row, column=10).value = day.get('max_min_temperature').get('max_temperature_night')
        sheet.cell(row=row, column=11).value = day.get('plesure')
        sheet.cell(row=row, column=12).value = day.get('humidity').get('humidity_morning')
        sheet.cell(row=row, column=13).value = day.get('humidity').get('humidity_day')
        sheet.cell(row=row, column=14).value = day.get('humidity').get('humidity_evening')
        sheet.cell(row=row, column=15).value = day.get('humidity').get('humidity_night')
        sheet.cell(row=row, column=16).value = day.get('condition').get('condition_morning')
        sheet.cell(row=row, column=17).value = day.get('condition').get('condition_day')
        sheet.cell(row=row, column=18).value = day.get('condition').get('condition_evening')
        sheet.cell(row=row, column=19).value = day.get('condition').get('condition_night')
        row+=1
    filename = f"weather_in_city_{city}_on_7_days.xlsx"
    if os.path.isdir('excel_docs') == False:
        os.mkdir("excel_docs")
    wb.save(f'excel_docs/{filename}')
    logger.info('save_in_excel')






    
    

